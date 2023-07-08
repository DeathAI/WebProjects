import xlsxwriter
import requests
from bs4 import BeautifulSoup
# import ptime
bal=10000.00
stockList=[]
stockName=[]
stockLinks=[]
accStock=[]

def price(url):
    st_response = requests.get(url)
    st_soup = BeautifulSoup(st_response.content, "html.parser")
    stock_div = st_soup.find("div", id="nsecp")
    if stock_div!=None:
        return stock_div.text.strip()
    else:
        return None

def printPrice():
    name=input('Enter the Name of Stock : ').lower().replace(' ','')
    if name in stockName:
        i=stockName.index(name)
        url = stockLinks[i]
        stPrice=price(url)
        if stPrice==None:
            print(name+"'s Data is not available")
        else:
            print("Price : ",stPrice)
    else:
        print(name+"'s Data is not available")

def displayList():
    i=1
    print('-------- All Stocks of India --------')
    for stock in stockList:
        print('\t',i,stock)
        i+=1

def INDlist():
    global stockLinks
    global stockList
    url = "https://www.moneycontrol.com/india/stockpricequote/"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, "html.parser")
    stocks = links = []
    for link in soup.find_all('a', class_='bl_12'):
        stock_name = link.get_text().strip()
        if stock_name=="" or stock_name==" ":
            continue
        stockList.append(stock_name)
        stockLinks.append(link['href'])

def accdetail():
    print(
         '''

-------------- Account --------------
Name : Rishi R Shah
i'd : 987263582
Phone Number : 9904547173
Balance : ''',bal,'''
-------------------------------------

            ''')

def GetDataExcel(file_path):
    import openpyxl
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    
    # Create an empty list to store the data
    data = []
    
    # Iterate through each row in the worksheet
    for row in worksheet.iter_rows(values_only=True):
        # Add the row to the data list
        data.append(list(row))
    return data

def writerows(file_path, data):
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet()
    for i, row in enumerate(data):
        worksheet.write_row(i, 0, row)
    workbook.close()

def addstock(name, quntity, stPrice):  
    global accStock
    import os
    import dates
    if not os.path.isfile('accStock.xlsx'):
        open('accStock.xlsx','w')
        workbook = xlsxwriter.Workbook('accStock.xlsx')
        worksheet = workbook.add_worksheet()
        # row=['Name', 'Quntity', 'Price']
        worksheet.write_row(0, 0, ['Name', 'Quntity', 'Price', 'Date'])
        workbook.close()
    data=GetDataExcel('accStock.xlsx')
    date=dates.datestr
    data.append([name, str(quntity), str(stPrice).replace(',',''), date])
    writerows('accStock.xlsx',data)
    accStock.append([name, str(quntity), str(stPrice).replace(',',''), date])

def buy():
    global bal
    name=input('Enter the Name of Stock : ').lower().replace(' ','')
    if name in stockName:
        i=stockName.index(name)
        url = stockLinks[i]
        stPrice=price(url)
        if stPrice==None:
            print(name+"'s Data is not available")
        else:
            quntity=int(input("How many stock you Want to buy : "))
            if bal < (float(stPrice.replace(',','')) * quntity):
                print('Sorry you can not purchse this stocks your balance is ',bal)
            else:
                print("Are you sure to buy",quntity,'stocks of ',name," at Price ",float(stPrice.replace(',','')) * quntity,' [Yes/No] : ',end='')
                ch=input()
                if ch[0].lower()=='y':
                    bal-=(float(stPrice.replace(',','')) * quntity)
                    print(quntity,' Stocks of ',name,' is successfully buyed by Rishi R Shah at cost of ',stPrice)
                    f=open('accinfo.txt','w')
                    f.write(str(bal))
                    addstock(name,quntity,stPrice)
                else:
                    print('Order has been censel sucessfully.')
    else:
        print(name+"'s Data is not available")

def removeStock(x):
    Data = GetDataExcel('accStock.xlsx')
    Data = Data[0]
    data = list(x)
    data.insert(0,Data)
    writerows('accStock.xlsx',data)

def sell():
    global bal,accStock
    name=input('Enter the Name of Stock : ').lower().replace(' ','')
    accStockNames = [x[0] for x in accStock]
    if name in accStockNames:
        no=0
        for AS in accStock:
            if AS[0]==name:
                no+=int(AS[1])
        print('You have',no,'Stocks of',stockList[stockName.index(name)])
        sellNo=int(input('so, How much stocks you want to sell : '))
        if sellNo>no:
            print('You has only',no,'Stocks of',stockList[stockName.index(name)])
        else:
            tmp=0
            i=0
            stPrice=price(stockLinks[stockName.index(name)])
            increment=sellNo*float(str(stPrice).replace(',',''))
            print('Are you sure to sell',sellNo,'of',name,'at price',increment,' [Yes/No] : ',end='')
            ch=input()
            if ch[0].lower()=='y':
                bal+=increment
                print(sellNo,' Stocks of ',name,' is successfully sold by Rishi R Shah at cost of ',stPrice)
            else:
                print('Selling process has been censel sucessfully.')
                return
            for AS in accStock:
                if sellNo==0:
                    break
                if AS[0]==name:
                    if int(AS[1])>sellNo:
                        tmp+=float(AS[2])*sellNo
                        AS[1]=str(int(AS[1])-sellNo)
                        sellNo=0
                    else:
                        tmp+=float(AS[1])*float(AS[2])
                        sellNo=sellNo-int(AS[1])
                        accStock.pop(accStock.index(AS))
                i+=1
            margin = increment - tmp
            if margin<=0:
                print('You sold your stock in',margin,"'s loss")
            else:
                print('You sold your stock in',margin,"'s profit")
            f=open('accinfo.txt','w')
            f.write(str(bal))
            removeStock(accStock)
    else:
        print("You have 0 stocks of",name)

def requireData():
    import socket
    global stockList, stockLinks, bal, accStock
    def is_connected():
        try:
            socket.create_connection(("8.8.8.8", 53))
            return True
        except OSError:
            pass
        return False
    if not is_connected():
        print("if possible then connecte your device to the internet.")
        f=open('stockList.txt','r')
        stockList=f.readlines()
        f=open('stockLinks.txt','r')
        stockLinks=f.readlines()
    else:
        print("Loadding ...")
        INDlist()
        f=open('stockList.txt','w')
        for stock in stockList:
            f.write(stock+"\n")
        f=open('stockLinks.txt','w')
        for link in stockLinks:
            f.write(link+"\n")
        f.close()
    try:
        f=open('accinfo.txt','r')
        bal=float(f.read())
    except:
        f=open('accinfo.txt','w')
        f.write('10000.0')
        f.close()
        f=open('accinfo.txt','r')
        bal=float(f.read())
    import os
    if os.path.isfile('accStock.xlsx'):
        accStock=GetDataExcel('accStock.xlsx')
        accStock.pop(0)
        for i in range(0,len(accStock)):
            accStock[i][1]=str(accStock[i][1]).replace(',','')
            accStock[i][2]=str(accStock[i][2]).replace(',','')
    for name in stockList:
        stockName.append(str(name).lower().replace(' ','')) 
requireData()

def printAccSotck():
    global accStock
    print('\n')
    print('---------------- Stocks ----------------')
    print('Stock', '', 'Quntity', 'Price', 'Date', sep='\t')
    print('----------------------------------------')
    for i in accStock:
        if len(stockList[stockName.index(i[0])])<5:
            print(stockList[stockName.index(i[0])], '', i[1], int(float(i[2].replace(',',''))), i[3], sep='\t')
        else:
            print(stockList[stockName.index(i[0])], i[1], int(float(i[2].replace(',',''))), i[3], sep='\t')
    print('----------------------------------------')

def saveBal():
    f=open('accinfo.txt','w')
    f.write(str(bal))

mainLoop=True
while mainLoop:
    print(
        '',
        '---------------- menu ---------------',
        '1. Enter 1 for see all names of stocks',
        '2. Enter 2 for get values of stock',
        '3. Enter 3 for see your balance',
        '4. Enter 4 for see your own stocks list',
        '5. Enter 5 for buy stock',
        '6. Enter 6 for sell stock',
        '7. Enter 7 for short sell stocks',
        '8. press 8 for Exit ',
        '-------------------------------------',
        sep='\n',
        end='\n',
    )
    match(input("Enter your option here : ").replace(' ','')):
        case '1': displayList()
        case '2': printPrice()
        case '3': accdetail()
        case '4': printAccSotck()
        case '5': buy()
        case '6': sell()
        case '7': printPrice()
        case '8':mainLoop=False
        case 'q':mainLoop=False
        case 'Q':mainLoop=False
        case '!!zerobal':bal=0
        case '!!bal10000':bal+=10000
        case '!!savebal':saveBal()