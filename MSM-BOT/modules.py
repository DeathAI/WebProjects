import xlsxwriter
import requests
from bs4 import BeautifulSoup
import discord
from discord.ext import commands

# import ptime
bal=5000.00
stockList=[]
stockName=[]
stockLinks=[]
accStock=[]

def price(url):
    st_response = requests.get(url)
    st_soup = BeautifulSoup(st_response.content, "html.parser")
    stock_div = st_soup.find("div", id="nsecp")
    if stock_div!=None:
        return stock_div.text.strip()
    else:
        return None

def givePrice(stock = None):
    if stock==None:
        stock = input('Enter the Name : ')
    stock=stock.lower().replace(' ','')
    if stock in stockName:
        i=stockName.index(stock)
        url = stockLinks[i]
        stPrice=price(url)
        if stPrice==None:
            return f"{stockList[i]} is available"
        else:
            return f"{stockList[i]} : {stPrice}"
    else:
        return None

def allStocks():
    return stockList

def INDlist():
    global stockLinks
    global stockList
    url = "https://www.moneycontrol.com/india/stockpricequote/"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, "html.parser")
    stocks = links = []
    for link in soup.find_all('a', class_='bl_12'):
        stock_name = link.get_text().strip()
        if stock_name=="" or stock_name==" ":
            continue
        stockList.append(stock_name)
        stockLinks.append(link['href'])

def accdetail(userName):
    bal=getBal(userName)
    return str(f'''

-------------- Account --------------
Name : {userName}
Balance : {bal} :money_with_wings:
-------------------------------------

            ''')

def GetDataExcel(file_path):
    import openpyxl
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    data = []
    for row in worksheet.iter_rows(values_only=True):
        data.append(list(row))
    return data

def writerows(file_path, data):
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet()
    for i, row in enumerate(data):
        worksheet.write_row(i, 0, row)
    workbook.close()

def addstock(file,name, quntity, stPrice,account):  
    global accStock
    import os
    import dates
    if not os.path.isfile(file):
        open(file,'w')
        workbook = xlsxwriter.Workbook(file)
        worksheet = workbook.add_worksheet()
        # row=['Name', 'Quntity', 'Price']
        worksheet.write_row(0, 0, ['Name', 'Quntity', 'Price', 'Date'])
        workbook.close()
    data=GetDataExcel(file)
    date=dates.datestr
    data.append([name, str(quntity), str(stPrice).replace(',',''), date])
    writerows(file,data)
    accStock.append([name, str(quntity), str(stPrice).replace(',',''), date])

def editAcc(name, bal="", checIn=""):
    import os
    file='Accounts.xlsx'
    if not os.path.isfile(file):
        open(file,'w')
        workbook = xlsxwriter.Workbook(file)
        worksheet = workbook.add_worksheet()
        # row=['Name', 'Balance', 'Last Checked In']
        worksheet.write_row(0, 0, ['Name', 'Balance', 'Last Checked In'])
        workbook.close()
        return
    data = GetDataExcel(file)
    i=0
    for x in data:
        if(x[0]==name):
            if(checIn=="" and bal!=""):
                data[i] = [x[0], bal, x[2]]
                writerows(file,data)
            else:
                data[i] = [x[0], x[1], checIn]
                writerows(file,data)
            return
        i+=1
    
    
def isUserExist(name):
    import os
    file='Accounts.xlsx'
    if not os.path.isfile(file):
        open(file,'w')
        workbook = xlsxwriter.Workbook(file)
        worksheet = workbook.add_worksheet()
        # row=['Name', 'Balance', 'Last Checked In']
        worksheet.write_row(0, 0, ['Name', 'Balance', 'Last Checked In'])
        workbook.close()
        return False
    data = GetDataExcel(file)
    i=0
    for x in data:
        if(x[0]==name):
            return True
        i+=1
    return False

def getBal(name):
    data = GetDataExcel('Accounts.xlsx')
    i=0
    for x in data:
        if(x[0]==name):
            return x[1]
        i+=1

def getLastTime(name):
    data = GetDataExcel('Accounts.xlsx')
    i=0
    for x in data:
        if(x[0]==name):
            return x[2]
        i+=1

def addAcc(name,bal):
    data = GetDataExcel('Accounts.xlsx')
    data.append([name,bal,''])
    writerows('Accounts.xlsx',data)
    

def buy(bal,name,quntity,account):
    name=name.lower().replace(' ','')
    quntity=int(quntity)
    if name in stockName:
        i=stockName.index(name)
        url = stockLinks[i]
        stPrice=price(url)
        if stPrice==None:
            return "This Stocks's Data is not available"
        else:
            if bal < (float(stPrice.replace(',','')) * quntity):
                return'Sorry you can not purchse this stocks'
            else:
                bal-=(float(stPrice.replace(',','')) * quntity)
                print(quntity,' Stocks of ',name,f' is successfully buyed by {account} at cost of ',stPrice)
                addstock(account+'_Stock.xlsx',name,quntity,stPrice,account)
                editAcc(account,bal)
                return f"You Buy {stockList[stockName.index(name)]}'s {quntity} stock at {stPrice} :money_with_wings:"
    else:
        return "This Stocks's Data is not available"

def removeStock(x,userName):
    fileName = userName+'_Stock.xlsx'
    Data = GetDataExcel(fileName)
    Data = Data[0]
    data = list(x)
    data.insert(0,Data)
    writerows(fileName,data)

def userbal(file):
    f = open(file,'r')
    return int(f.read())

def getAccStock(userName):
    Data = GetDataExcel(userName+'_Stock.xlsx')
    Data.pop(0)
    return Data



def sell(bal,name,quntity,userName):
    accStock = getAccStock(userName)
    name=name.lower().replace(' ','')
    accStockNames = [x[0] for x in accStock]
    if name in accStockNames:
        no=0
        for AS in accStock:
            if AS[0]==name:
                no+=int(AS[1])
        sellNo=int(quntity)
        if sellNo>no:
            return f'You has only {no} Stocks of {stockList[stockName.index(name)]}'
        else:
            tmp=0
            i=0
            stPrice=price(stockLinks[stockName.index(name)])
            increment=sellNo*float(str(stPrice).replace(',',''))
            
            bal+=increment
            print(sellNo,' Stocks of ',stockList[stockName.index(name)],' is successfully sold by ',userName,' at cost of ',stPrice)

            for AS in accStock:
                if sellNo==0:
                    break
                if AS[0]==name:
                    if int(AS[1])>sellNo:
                        tmp+=float(AS[2])*sellNo
                        AS[1]=str(int(AS[1])-sellNo)
                        sellNo=0
                    else:
                        tmp+=float(AS[1])*float(AS[2])
                        sellNo=sellNo-int(AS[1])
                        accStock.pop(accStock.index(AS))
                i+=1
            margin = increment - tmp
            editAcc(userName,bal)
            removeStock(accStock,userName)
            if margin<=0:
                return f"You sold your {stockList[stockName.index(name)]} in {margin}'s loss"
            else:
                return f"You sold your {stockList[stockName.index(name)]} in {margin}'s profit"
    else:
        print("You have 0 stocks of",name)

def requireData():
    import socket
    global stockList, stockLinks, bal, accStock
    def is_connected():
        try:
            socket.create_connection(("8.8.8.8", 53))
            return True
        except OSError:
            pass
        return False
    if not is_connected():
        print("if possible then connecte your device to the internet.")
        f=open('stockList.txt','r')
        stockList=f.readlines()
        f=open('stockLinks.txt','r')
        stockLinks=f.readlines()
    else:
        print("Loadding ...")
        INDlist()
        f=open('stockList.txt','w')
        for stock in stockList:
            f.write(stock+"\n")
        f=open('stockLinks.txt','w')
        for link in stockLinks:
            f.write(link+"\n")
    isUserExist("")
    for name in stockList:
        stockName.append(str(name).lower().replace(' ','')) 

def printAccStock(name):
    from PIL import Image, ImageDraw, ImageFont

    accStock = getAccStock(name)
    font_size = 12
    font = ImageFont.truetype("arial.ttf", font_size)

    # Calculate the text dimensions for positioning
    text_height = font.getbbox("Sample Text")[3] - font.getbbox("Sample Text")[1]

    text_widths = []
    for i in accStock:
        stock_name = stockList[stockName.index(i[0])]
        text_width = font.getbbox(stock_name)[2] - font.getbbox(stock_name)[0]
        text_widths.append(text_width)

    # Calculate the maximum width for each column
    max_width_1 = max(text_widths)
    max_width_2 = font.getbbox("Quantity")[2] - font.getbbox("Quantity")[0]
    max_width_3 = font.getbbox("Price")[2] - font.getbbox("Price")[0]
    max_width_4 = font.getbbox("Date")[2] - font.getbbox("Date")[0]

    # Calculate the width and height of the image
    padding = 10
    column_padding = 20
    width = max_width_1 + max_width_2 + max_width_3 + max_width_4 + (4 * column_padding) + (2 * padding)
    height = (len(accStock) + 2) * text_height + (2 * padding)

    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)

    # Set the initial position for drawing
    x = padding
    y = padding

    # Draw the table headers
    draw.text((x, y), "Stock", fill="black", font=font)
    x += max_width_1 + column_padding

    draw.text((x, y), "Quantity", fill="black", font=font)
    x += max_width_2 + column_padding

    draw.text((x, y), "Price", fill="black", font=font)
    x += max_width_3 + column_padding

    draw.text((x, y), "Date", fill="black", font=font)

    # Draw the table rows
    y += text_height

    for i in accStock:
        stock_name = stockList[stockName.index(i[0])]

        x = padding
        draw.text((x, y), stock_name, fill="black", font=font)
        x += max_width_1 + column_padding

        draw.text((x, y), str(i[1]), fill="black", font=font)
        x += max_width_2 + column_padding

        draw.text((x, y), str(int(float(i[2].replace(' ,', '')))), fill="black", font=font)
        x += max_width_3 + column_padding

        draw.text((x, y), i[3], fill="black", font=font)

        y += text_height

    
    # Save the image with higher quality
    image.save("temp.jpg")
