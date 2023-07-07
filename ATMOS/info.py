# Private DB module for handling ExcelSheets
# Importing private Dates Module to use calender related Functions
import dates
def read_DB(filename):
    import openpyxl

    # Load the Excel file
    workbook = openpyxl.load_workbook(filename+'.xlsx')
    worksheet = workbook.active

    # Create an empty list to store the data
    data = []
    val=[[],[]]
    
    # Iterate over the rows in the worksheet
    for row in worksheet.iter_rows(values_only=True):
        # Append the row to the data list
        data.append(list(row))
    data.pop(0)

    # Print the data
    for row in data:
        val[0].append(row[0])
    for row in data:
        val[1].append(row[1])
    return val

def cls(filename):
    import os
    if os.path.isfile(filename):
        os.remove(filename)

def excelbody(file_name,month,year,list1,list2):
    import xlsxwriter

    workbook = xlsxwriter.Workbook(file_name)
    worksheet = workbook.add_worksheet()

    worksheet.write('A1', "I'd")
    worksheet.write('B1', "Name")
    
    row = 0
    col = 2
    for date in dates.getDates(month,year):
        worksheet.write(row, col, f'{date}-{month}-{year}')
        col+= 1
    row=1
    col=1
    for item in list1:
        worksheet.write(row,col,item)
        row+=1
    row=1
    col=0
    for item in list2:
        worksheet.write(row,col,item)
        row+=1
    # print(list1,list2)
    workbook.close()

def writecol(filename, attendance, pos, row):
        from openpyxl import load_workbook

        try:
            workbook = load_workbook(filename)
        except FileNotFoundError:
            workbook = load_workbook.Workbook()
            workbook.save(filename)

        try:
            worksheet = workbook.active
        except KeyError:
            worksheet = workbook.create_sheet(title='Sheet1')

        for i in range(0, len(attendance)):
            worksheet.cell(row=i+2, column=pos+1, value=attendance[i])
            print(i, pos, attendance[i - 1])
        workbook.save(filename)

def srhExcel(filename, attendance, data, date,row=None):
    # print(data,date)
    for i in range(0, len(data)):
        if  (data[i] == date):
            writecol(filename, attendance, i ,row)
            return

def excel_data(filename, attendance, date):
     
    import openpyxl
    
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active

    data = []
    
    for row in worksheet.iter_rows(values_only=True,max_col=33,min_row=1):
        data.append(list(row))
        
    date=str(date).replace(',','-').replace(' ','')[1:-1]
    if date not in data[0]:
        return
    
    filename=str(filename)
    if filename.endswith('.xlsx'):
        filename.replace('.xlsx','')
    srhExcel(filename, attendance, data[0], date)

import xlsxwriter

def writerows(file_path, data):
    # Create a new workbook and add a worksheet
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet()
    # Write each row of data to the worksheet
    for i, row in enumerate(data):
        worksheet.write_row(i, 0, row)
    # Close the workbook
    workbook.close()

# file_path = 'file.xlsx'
# data = [
#     ['Name', 'Age', 'City'],
#     ['John', 30, 'New York'],
#     ['Jane', 25, 'Los Angeles'],
#     ['Bob', 40, 'Chicago']
# ]
# writerows(file_path, data)

def del_row(file_path, id):
    data=GetDataExcel(file_path)
    ids = [x[0] for x in data]

    if id not in ids:
        return
    
    i=ids.index(id)
    data.pop(i)

    cls(file_path)
    writerows(file_path, data)

def GetDataExcel(file_path):
    import openpyxl
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    
    # Create an empty list to store the data
    data = []
    
    # Iterate through each row in the worksheet
    for row in worksheet.iter_rows(values_only=True):
        # Add the row to the data list
        data.append(row)
    return data
    

def add_xl(filename,id,name,val=None):
    counter=1
    data=GetDataExcel(filename)
    for i in data:
        if data[0]==i:
            continue
        if id<i[0]:
            data.insert(counter,[id,name])
            cls(filename)
            writerows(filename,data)
            return
        counter+=1
    data.insert(counter,[id,name])
    cls(filename)
    writerows(filename,data)
    if val==None:
        for i in range(1,dates.todaydt[0]+1):
            srhExcel(filename,False,i+2,counter)
    return

def edit_xl(filename,id1,id2,name1,name2):
    x=GetDataExcel(filename)
    data=[]
    for val in x:
        data.append(list(val))
    for i in data:
        if id1==i[0] and name1==i[1]:
            i[0],i[1]=id2,name2
    writerows(filename,data)

def report(ipFile,opFile,id,month,year):
    x=GetDataExcel(ipFile)
    ids = [int(x[j][0]) for j in range(1,len(x))]
    data=[]
    c=0
    if id in ids:
        for val in x:
            if c==0 or val[0]==id:
                data.append(list(val))
            c+=1
        data[0].append('persentage')
        per=round(float(data[1].count(True)*100)/len(dates.getDates(month,year)),2)
        print(per,data[1].count(True),len(dates.getDates(month,year)))
        data[1].append(f'{per} %')
        writerows(opFile,data)

def reportGroup(ipFile,opFile,month,year):
    x=GetDataExcel(ipFile)
    ids = [int(x[j][0]) for j in range(1,len(x))]
    data=[]
    for val in x:
        data.append(list(val))
    data[0].append('persentage')
    for c in range(1,len(data)):
        per=round(float(data[c].count(True)*100)/len(dates.getDates(month,year)),2)
        print(per,data[c].count(True),len(dates.getDates(month,year)))
        data[c].append(f'{per} %')
    print(data)
    writerows(opFile,data)

    # print(ids)
def rg(ipFile):
    reportGroup(ipFile,'report.xlsx',dates.month,dates.year)
    GetDataExcel('report.xlsx')
# reportGroup('names.xlsx','areport.xlsx',3,2023)
# rg('soome.xlsx')
# edit_xl('names.xlsx',1,4,'Rishi','Tanish')

# add_xl('names.xlsx',4,'Tanish')
# add_xl('names.xlsx',0,'Nikunj')
    
# print(GetDataExcel("soome.xlsx"))
#excelbody('names.xlsx',3,2023,['Rishi','Omi','Aditya'],[1,2,3])

# del_row('names.xlsx', 5511)     # <== This is Testing Funtion Uncomment For testing Delete Attendance Data for persons
# for i in range(1,32):
# excel_data('names.xlsx', [False,False,False], [22,3,2023])     # <== This is Testing Funtion Uncomment For testing Insert Attendance Data for perticuler Date