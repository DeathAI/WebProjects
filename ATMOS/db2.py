# Private DB module for handling ExcelSheets
# Importing private Dates Module to use calender related Functions
import dates
def read_DB(filename):
    import openpyxl
    
    # Load the Excel file
    workbook = openpyxl.load_workbook(filename+".xlsx")
    worksheet = workbook.active
    print(filename)
    # Create an empty list to store the data
    data = []
    val=[[],[]]
    
    # Iterate over the rows in the worksheet
    for row in worksheet.iter_rows(values_only=True):
        # Append the row to the data list
        data.append(list(row))
    data.pop(0)

    # Print the data
    for row in data:
        val[0].append(row[0])
    for row in data:
        val[1].append(row[1])
    return val

def cls(filename):
    import os
    os.remove(filename+'.xlsx')

def edit(filename, data ,pos):
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(filename)
    except FileNotFoundError:
        workbook = load_workbook.Workbook()
        workbook.save(filename)

    try:
        worksheet = workbook.active
    except KeyError:
        worksheet = workbook.create_sheet(title='Sheet1')

    for i in range(0, len(data)):
        worksheet.cell(row=i+2, column=pos+1, value=data[i])
        worksheet.cell(row=i+2, column=pos+1, value=data[i])
        print(i, pos, data[i - 1])
    workbook.save(filename)

def add():
    workbook = ("kaipn.xlsx")

    # Access the first worksheet in the Excel file
    worksheet = workbook.getWorksheets().get(0)

    # Insert a row into the worksheet at 3rd position
    worksheet.getCells().insertRows(2,1)

    # Save the modified Excel file in default (that is Excel 2003) format
    workbook.save("kaipn.xlsx")

def update(file,name1,id1,name2,id2):
    
    import openpyxl
    file
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.active

    data = []
    
    for row in worksheet.iter_rows(values_only=True,max_col=33,min_row=1):
        data.append(list(row))

    print("flag 1")
    i=0
    while i != len(data):
        print(data)
        if  (data[i]=="Name"):
            print(i)
            edit(file, id2, i)
            if  ("ID" == data[i]):
                edit(file, name2, i)
        i+=1
        return

def delete():
    pass


def excelbody(file_name,month,year,list1,list2):
    import xlsxwriter

    workbook = xlsxwriter.Workbook(file_name)
    worksheet = workbook.add_worksheet()

    worksheet.write('A1', "I'd")
    worksheet.write('B1', "Name")
    
    row = 0
    col = 2
    for date in dates.getDates(month,year):
        worksheet.write(row, col, f'{date}-{month}-{year}')
        col+= 1
    row=1
    col=1
    for item in list1:
        worksheet.write(row,col,item)
        row+=1
    row=1
    col=0
    for item in list2:
        worksheet.write(row,col,item)
        row+=1
    print(list1,list2)
    workbook.close()

def srhExcel(filename, attendance, data, date):
    def writecol(filename, attendance, pos):
        from openpyxl import load_workbook

        try:
            workbook = load_workbook(filename)
        except FileNotFoundError:
            workbook = load_workbook.Workbook()
            workbook.save(filename)

        try:
            worksheet = workbook.active
        except KeyError:
            worksheet = workbook.create_sheet(title='Sheet1')

        for i in range(0, len(attendance)):
            worksheet.cell(row=i+2, column=pos+1, value=attendance[i])
            print(i, pos, attendance[i - 1])
        workbook.save(filename)

    
    for i in range(0, len(data)):
        if  (data[i] == date):
            writecol(filename, attendance, i)
            return

def excel_data(filename, attendance, date):
    
    import openpyxl
    
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active

    data = []
    
    for row in worksheet.iter_rows(values_only=True,max_col=33,min_row=1):
        data.append(list(row))


    date=str(date).replace(',','/').replace(' ','')[1:-1]

    print("flag 1")
    print(date)
    print(data[0])
    if date not in data[0]:
        return
    print("flag 2")
    filename=str(filename)
    if filename.endswith('.xlsx'):
        filename.replace('.xlsx','')
    print(filename)
    srhExcel(filename, attendance, data[0], date)



#excel_data('kaipn.xlsx', [False,True,False], [23,3,2023])     # <== This is Testing Funtion Uncomment For testing Insert Attendance Data for perticuler Date
#update("kaipn.xlsx","Tanish",23,"omi",35)