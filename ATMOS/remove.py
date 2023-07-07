# Private DB module for handling ExcelSheets
# Importing private Dates Module to use calender related Functions
import dates
def read_DB(filename):
    import openpyxl

    # Load the Excel file
    workbook = openpyxl.load_workbook(filename+'.xlsx')
    worksheet = workbook.active

    # Create an empty list to store the data
    data = []
    val=[[],[]]
    
    # Iterate over the rows in the worksheet
    for row in worksheet.iter_rows(values_only=True):
        # Append the row to the data list
        data.append(list(row))
    data.pop(0)

    # Print the data
    for row in data:
        val[0].append(row[0])
    for row in data:
        val[1].append(row[1])
    return val

def cls(filename):
    import os
    if os.path.isfile(filename):
        os.remove(filename)

def excelbody(file_name,month,year,list1,list2):
    import xlsxwriter

    workbook = xlsxwriter.Workbook(file_name)
    worksheet = workbook.add_worksheet()

    worksheet.write('A1', "I'd")
    worksheet.write('B1', "Name")
    
    row = 0
    col = 2
    for date in dates.getDates(month,year):
        worksheet.write(row, col, f'{date}-{month}-{year}')
        col+= 1
    row=1
    col=1
    for item in list1:
        worksheet.write(row,col,item)
        row+=1
    row=1
    col=0
    for item in list2:
        worksheet.write(row,col,item)
        row+=1
    # print(list1,list2)
    workbook.close()

def srhExcel(filename, attendance, data, date):
    def writecol(filename, attendance, pos):
        from openpyxl import load_workbook

        try:
            workbook = load_workbook(filename)
        except FileNotFoundError:
            workbook = load_workbook.Workbook()
            workbook.save(filename)

        try:
            worksheet = workbook.active
        except KeyError:
            worksheet = workbook.create_sheet(title='Sheet1')

        for i in range(0, len(attendance)):
            worksheet.cell(row=i+2, column=pos+1, value=attendance[i])
            print(i, pos, attendance[i - 1])
        workbook.save(filename)

    print(data,date)
    for i in range(0, len(data)):
        if  (data[i] == date):
            writecol(filename, attendance, i)
            return

def excel_data(filename, attendance, date):
     
    import openpyxl
    
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active

    data = []
    
    for row in worksheet.iter_rows(values_only=True,max_col=33,min_row=1):
        data.append(list(row))
        
    date=str(date).replace(',','-').replace(' ','')[1:-1]
    if date not in data[0]:
        return
    
    filename=str(filename)
    if filename.endswith('.xlsx'):
        filename.replace('.xlsx','')
    srhExcel(filename, attendance, data[0], date)



import xlsxwriter

def writerows(file_path, data):
    # Create a new workbook and add a worksheet
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet()

    # Write each row of data to the worksheet
    for i, row in enumerate(data):
        worksheet.write_row(i, 0, row)

    # Close the workbook
    workbook.close()

# file_path = 'file.xlsx'
# data = [
#     ['Name', 'Age', 'City'],
#     ['John', 30, 'New York'],
#     ['Jane', 25, 'Los Angeles'],
#     ['Bob', 40, 'Chicago']
# ]
# writerows(file_path, data)
def del_row(file_path, id):
    import openpyxl
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    
    # Create an empty list to store the data
    data = []
    
    # Iterate through each row in the worksheet
    for row in worksheet.iter_rows(values_only=True):
        # Add the row to the data list
        data.append(row)

    ids = [x[0] for x in data]

    
    if id not in ids:
        return
    
    i=ids.index(id)
    data.pop(i)

    cls(file_path)
    writerows(file_path, data)

# excelbody('soome.xlsx',3,2023,['Rishi','Omi','Aditya'],[1,2,3])

# del_row('names.xlsx', 5511)     # <== This is Testing Funtion Uncomment For testing Insert Attendance Data for perticuler Date

# excel_data('names.xlsx', [True], [31,3,2023])     # <== This is Testing Funtion Uncomment For testing Insert Attendance Data for perticuler Date